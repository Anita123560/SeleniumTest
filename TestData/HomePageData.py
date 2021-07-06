import openpyxl


class HomePageData:

    test_homepage_data = [{"firstname":"Anita", "lastname":"S", "password":"Pass123", "gender":"Male"},
                            {"firstname":"Anita1", "lastname":"S1", "password":"Pass1231", "gender":"Female"}]
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("E:\\Selenium_Python\\Rahul Shetty training code\\ExcelTestData\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

